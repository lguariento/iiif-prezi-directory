from iiif_prezi.factory import ManifestFactory
from openpyxl import load_workbook
import os
import warnings
warnings.simplefilter("ignore")

spreadsheet = load_workbook(filename="McLagan_MSGen1042.xlsx", read_only=True)
sheet = spreadsheet.active

manuscripts = ["MS_Gen_1042_1", "MS_Gen_1042_10", "MS_Gen_1042_100", "MS_Gen_1042_101", "MS_Gen_1042_102", "MS_Gen_1042_103", "MS_Gen_1042_104", "MS_Gen_1042_105", "MS_Gen_1042_106a", "MS_Gen_1042_106b", "MS_Gen_1042_107", "MS_Gen_1042_108", "MS_Gen_1042_109", "MS_Gen_1042_11", "MS_Gen_1042_110", "MS_Gen_1042_111", "MS_Gen_1042_112", "MS_Gen_1042_113", "MS_Gen_1042_114", "MS_Gen_1042_115", "MS_Gen_1042_116", "MS_Gen_1042_117", "MS_Gen_1042_118", "MS_Gen_1042_119", "MS_Gen_1042_12", "MS_Gen_1042_120", "MS_Gen_1042_121", "MS_Gen_1042_122", "MS_Gen_1042_123", "MS_Gen_1042_124", "MS_Gen_1042_125", "MS_Gen_1042_126", "MS_Gen_1042_127", "MS_Gen_1042_128", "MS_Gen_1042_129a", "MS_Gen_1042_129b", "MS_Gen_1042_13", "MS_Gen_1042_130", "MS_Gen_1042_132", "MS_Gen_1042_133", "MS_Gen_1042_134", "MS_Gen_1042_135a", "MS_Gen_1042_135b", "MS_Gen_1042_136", "MS_Gen_1042_137", "MS_Gen_1042_138", "MS_Gen_1042_139", "MS_Gen_1042_14", "MS_Gen_1042_140", "MS_Gen_1042_141", "MS_Gen_1042_142", "MS_Gen_1042_143", "MS_Gen_1042_144", "MS_Gen_1042_145", "MS_Gen_1042_146", "MS_Gen_1042_147", "MS_Gen_1042_148", "MS_Gen_1042_149", "MS_Gen_1042_15", "MS_Gen_1042_150", "MS_Gen_1042_151", "MS_Gen_1042_152", "MS_Gen_1042_153", "MS_Gen_1042_154", "MS_Gen_1042_155", "MS_Gen_1042_156", "MS_Gen_1042_157", "MS_Gen_1042_158", "MS_Gen_1042_159", "MS_Gen_1042_16", "MS_Gen_1042_160", "MS_Gen_1042_161", "MS_Gen_1042_162", "MS_Gen_1042_163", "MS_Gen_1042_164", "MS_Gen_1042_165a", "MS_Gen_1042_165b", "MS_Gen_1042_166", "MS_Gen_1042_167", "MS_Gen_1042_168", "MS_Gen_1042_169", "MS_Gen_1042_17", "MS_Gen_1042_170", "MS_Gen_1042_171", "MS_Gen_1042_172", "MS_Gen_1042_173", "MS_Gen_1042_174", "MS_Gen_1042_175", "MS_Gen_1042_176", "MS_Gen_1042_177", "MS_Gen_1042_178", "MS_Gen_1042_179", "MS_Gen_1042_18", "MS_Gen_1042_180", "MS_Gen_1042_181", "MS_Gen_1042_182", "MS_Gen_1042_183", "MS_Gen_1042_184", "MS_Gen_1042_185", "MS_Gen_1042_186", "MS_Gen_1042_187", "MS_Gen_1042_188", "MS_Gen_1042_189", "MS_Gen_1042_19", "MS_Gen_1042_190", "MS_Gen_1042_191", "MS_Gen_1042_192", "MS_Gen_1042_193", "MS_Gen_1042_194", "MS_Gen_1042_195", "MS_Gen_1042_196", "MS_Gen_1042_197", "MS_Gen_1042_198", "MS_Gen_1042_199", "MS_Gen_1042_2", "MS_Gen_1042_20", "MS_Gen_1042_200", "MS_Gen_1042_201", "MS_Gen_1042_203", "MS_Gen_1042_204", "MS_Gen_1042_205", "MS_Gen_1042_206", "MS_Gen_1042_207", "MS_Gen_1042_208", "MS_Gen_1042_209", "MS_Gen_1042_21", "MS_Gen_1042_210", "MS_Gen_1042_211", "MS_Gen_1042_212", "MS_Gen_1042_213", "MS_Gen_1042_214", "MS_Gen_1042_216", "MS_Gen_1042_217", "MS_Gen_1042_218", "MS_Gen_1042_219", "MS_Gen_1042_22", "MS_Gen_1042_220", "MS_Gen_1042_221", "MS_Gen_1042_222a", "MS_Gen_1042_222b", "MS_Gen_1042_222c", "MS_Gen_1042_223", "MS_Gen_1042_224", "MS_Gen_1042_225", "MS_Gen_1042_226", "MS_Gen_1042_227", "MS_Gen_1042_228", "MS_Gen_1042_229", "MS_Gen_1042_23", "MS_Gen_1042_230", "MS_Gen_1042_231", "MS_Gen_1042_232", "MS_Gen_1042_233", "MS_Gen_1042_234", "MS_Gen_1042_235", "MS_Gen_1042_236", "MS_Gen_1042_237", "MS_Gen_1042_238", "MS_Gen_1042_239", "MS_Gen_1042_24", "MS_Gen_1042_240", "MS_Gen_1042_241", "MS_Gen_1042_242", "MS_Gen_1042_243", "MS_Gen_1042_244", "MS_Gen_1042_245", "MS_Gen_1042_246", "MS_Gen_1042_247", "MS_Gen_1042_248", "MS_Gen_1042_249", "MS_Gen_1042_24a", "MS_Gen_1042_25", "MS_Gen_1042_250", "MS_Gen_1042_251", "MS_Gen_1042_252", "MS_Gen_1042_253", "MS_Gen_1042_254", "MS_Gen_1042_26", "MS_Gen_1042_27", "MS_Gen_1042_28", "MS_Gen_1042_29", "MS_Gen_1042_3", "MS_Gen_1042_30", "MS_Gen_1042_31", "MS_Gen_1042_32", "MS_Gen_1042_33", "MS_Gen_1042_34", "MS_Gen_1042_35", "MS_Gen_1042_36", "MS_Gen_1042_37", "MS_Gen_1042_38", "MS_Gen_1042_39", "MS_Gen_1042_4", "MS_Gen_1042_40", "MS_Gen_1042_41", "MS_Gen_1042_42", "MS_Gen_1042_43", "MS_Gen_1042_44", "MS_Gen_1042_45", "MS_Gen_1042_47", "MS_Gen_1042_48", "MS_Gen_1042_49", "MS_Gen_1042_5", "MS_Gen_1042_50", "MS_Gen_1042_51", "MS_Gen_1042_52", "MS_Gen_1042_53", "MS_Gen_1042_54", "MS_Gen_1042_54a", "MS_Gen_1042_55", "MS_Gen_1042_56", "MS_Gen_1042_57", "MS_Gen_1042_58", "MS_Gen_1042_59", "MS_Gen_1042_60", "MS_Gen_1042_61", "MS_Gen_1042_62", "MS_Gen_1042_63", "MS_Gen_1042_64", "MS_Gen_1042_65", "MS_Gen_1042_66", "MS_Gen_1042_67", "MS_Gen_1042_68", "MS_Gen_1042_69", "MS_Gen_1042_7", "MS_Gen_1042_70", "MS_Gen_1042_71", "MS_Gen_1042_72", "MS_Gen_1042_73", "MS_Gen_1042_74", "MS_Gen_1042_75", "MS_Gen_1042_76", "MS_Gen_1042_77", "MS_Gen_1042_78", "MS_Gen_1042_79", "MS_Gen_1042_79a", "MS_Gen_1042_8", "MS_Gen_1042_80", "MS_Gen_1042_81", "MS_Gen_1042_81a", "MS_Gen_1042_82", "MS_Gen_1042_83", "MS_Gen_1042_84", "MS_Gen_1042_85", "MS_Gen_1042_86", "MS_Gen_1042_87", "MS_Gen_1042_88", "MS_Gen_1042_89", "MS_Gen_1042_9", "MS_Gen_1042_90", "MS_Gen_1042_91", "MS_Gen_1042_92", "MS_Gen_1042_93", "MS_Gen_1042_94", "MS_Gen_1042_95", "MS_Gen_1042_96", "MS_Gen_1042_97", "MS_Gen_1042_98", "MS_Gen_1042_99"]

ms_with_ranges =["MS Gen 1042/10","MS Gen 1042/102","MS Gen 1042/105","MS Gen 1042/106a","MS Gen 1042/106b","MS Gen 1042/109","MS Gen 1042/111","MS Gen 1042/114","MS Gen 1042/115","MS Gen 1042/118","MS Gen 1042/120","MS Gen 1042/122","MS Gen 1042/125","MS Gen 1042/126","MS Gen 1042/129a","MS Gen 1042/129b","MS Gen 1042/13","MS Gen 1042/130","MS Gen 1042/132","MS Gen 1042/135b","MS Gen 1042/136","MS Gen 1042/137","MS Gen 1042/139","MS Gen 1042/14","MS Gen 1042/140","MS Gen 1042/141","MS Gen 1042/142","MS Gen 1042/143","MS Gen 1042/145","MS Gen 1042/146","MS Gen 1042/148","MS Gen 1042/150","MS Gen 1042/151","MS Gen 1042/153","MS Gen 1042/154","MS Gen 1042/156","MS Gen 1042/160","MS Gen 1042/161","MS Gen 1042/162","MS Gen 1042/163","MS Gen 1042/165b","MS Gen 1042/166","MS Gen 1042/167","MS Gen 1042/168","MS Gen 1042/169","MS Gen 1042/170","MS Gen 1042/177","MS Gen 1042/18","MS Gen 1042/180","MS Gen 1042/181","MS Gen 1042/184","MS Gen 1042/185","MS Gen 1042/186","MS Gen 1042/187","MS Gen 1042/19","MS Gen 1042/190","MS Gen 1042/192","MS Gen 1042/193","MS Gen 1042/194","MS Gen 1042/195","MS Gen 1042/199","MS Gen 1042/2","MS Gen 1042/20","MS Gen 1042/200","MS Gen 1042/201","MS Gen 1042/204","MS Gen 1042/205","MS Gen 1042/209","MS Gen 1042/210","MS Gen 1042/211","MS Gen 1042/213","MS Gen 1042/216","MS Gen 1042/219","MS Gen 1042/22","MS Gen 1042/222a","MS Gen 1042/222b","MS Gen 1042/222c","MS Gen 1042/225","MS Gen 1042/226","MS Gen 1042/227","MS Gen 1042/229","MS Gen 1042/23","MS Gen 1042/230","MS Gen 1042/233","MS Gen 1042/235","MS Gen 1042/239","MS Gen 1042/240","MS Gen 1042/241","MS Gen 1042/244","MS Gen 1042/25","MS Gen 1042/26","MS Gen 1042/27","MS Gen 1042/29","MS Gen 1042/3","MS Gen 1042/33","MS Gen 1042/36","MS Gen 1042/39","MS Gen 1042/4","MS Gen 1042/45","MS Gen 1042/47","MS Gen 1042/5","MS Gen 1042/50","MS Gen 1042/51","MS Gen 1042/52","MS Gen 1042/53","MS Gen 1042/54","MS Gen 1042/58","MS Gen 1042/59","MS Gen 1042/61","MS Gen 1042/62","MS Gen 1042/64","MS Gen 1042/67","MS Gen 1042/68","MS Gen 1042/69","MS Gen 1042/70","MS Gen 1042/73","MS Gen 1042/76","MS Gen 1042/77","MS Gen 1042/80","MS Gen 1042/81","MS Gen 1042/82","MS Gen 1042/83","MS Gen 1042/85","MS Gen 1042/87","MS Gen 1042/89","MS Gen 1042/9","MS Gen 1042/90","MS Gen 1042/91","MS Gen 1042/94","MS Gen 1042/96","MS Gen 1042/97","MS Gen 1042/98","MS Gen 1042/99"]

prezi_dir = "/tmp"

fac = ManifestFactory()
fac.set_debug("error")
fac.set_iiif_image_info()
fac.set_base_prezi_uri("https://iiif.gla.ac.uk/")
fac.set_base_prezi_dir(prezi_dir)

ranges_done = []
manifests = []

def save_manifest(ms):
    data = manifest.toString(compact=False)
    fh = open("manifests/" + ms[1] + '.json', mode="w", encoding="utf-8")
    fh.write(data)
    fh.close()
    print('Manifest for ' + ms[0] + ' saved')

for ms in sheet.iter_rows(min_row=3, values_only=True):
    if (ms[2] == None) and (ms[0] not in ranges_done): # Check that it's not a range row.
        manifests.append(ms[0]) # Add it to the list of created manifests
        title = ms[0]
        manifest = fac.manifest(ident="manifest/" + ms[1] + ".json", label=title)
        manifest.set_metadata({"Physical description": ms[4], 'Bibliography': ms[8], 'Library record ID': str(ms[9])})
        manifest.attribution = "<span>Photo: © Archives & Special Collections, University of Glasgow Library. Terms of use: <a href=\"https://creativecommons.org/licenses/by-nc/4.0/\">CC-BY-NC</a></span>"
        if ms[7] != None:
            manifest.set_metadata({"Commentary": ms[7]})
        manifest.description = ms[6]
        seq = manifest.sequence(ident=ms[1] + ".json", label='Current page order')
        image_dir = "/Users/luca/Documents/development/iiif-prezi/images/" + ms[1] + "/images/"
        fac.set_base_image_dir(image_dir)
        fac.set_base_image_uri("https://iiif.gla.ac.uk/iiif/" + ms[1] + "/images")
        print('Getting data for ' + ms[0])
        for fn in sorted(os.listdir(image_dir)):
            if fn.endswith(".jp2"): # Makes sure that it picks up only images
                ident = fn[:-4]
                cvs_title = ident.replace("_", " ").title().split(' ')[-1]
                cvs = seq.canvas(ident=ident, label=cvs_title)
                #cvs.add_image_annotation(fn, True)

                anno = cvs.annotation(ident=ident + ".json", label=title)
                image = anno.image(ident=fn, iiif=True)
                image.set_hw_from_iiif()
                cvs.set_hw(image.height, image.width)

        if ms[0] not in ms_with_ranges: # If it's not a ms with ranges add the title and save the manifest
            if ms[5] != None:
                manifest.set_metadata({"Title": ms[5]}) # Add the title of the poem if not null

            save_manifest(ms)

        else: # ... else it has ranges defined, which we'll take care of.
            for poem in sheet.iter_rows(min_row=3, values_only=True): # Loops again through all the cells...
                if (poem[0] == ms[0]) and (poem[3] != None): # ... and find the ones belonging to the manuscript we're building the manifest for, and which have something in the 'range folios' cell.
                    print('found poem for ' + poem[1] + '(' + poem[2] + ')')
                    label = poem[5] if poem[5]!= None else poem[2] # If the title of the poem is empty then label the tange with the fol_ ... data.
                    rng = manifest.range(ident=poem[1] + '/' + poem[2], label=label)
                    if poem[6] != None:
                        rng.set_metadata({"Description": poem[6]})
                    if ',' in str(poem[3]): # Checks if there's more than one value, i.e. more than one page in this range.
                        for canvas in poem[3].split(','): # If so, loops through these comma-separated values.
                            rng.add_canvas(seq.canvases[int(canvas) - 1])
                            print('Added folio ' + canvas + ' to range ' + label)
                    else: # If not add the only page that there is.
                        rng.add_canvas(seq.canvases[int(poem[3]) - 1])
                    ranges_done.append(ms[0]) # Append this manuscript to the ranges_done list so that we avoid being picked up by the first loop.

            save_manifest(ms)


    elif ms[0] in ranges_done:
        pass
